#!/usr/bin/env python3
"""report_build.py — entries JSON + last month's workbook -> this month's report.

    ./report_build.py --entries e.json --template "…_Jun'26.xlsx" \
                      --out "…_Jul'26.xlsx" --month 2026-07 --rate 120 --bt keep

Rebuilds the "Details" tab from the time entries, grouped by the "[Prefix]" that
leads a description, and writes each group's money into the matching project row
on the summary tab. Layout per group:

    <group header band>
    entries…
    Subtotal - hours          <- only when the group also carries a business trip
    Business trip             <- preserved verbatim from the source workbook
    Subtotal - <PR-code>
    (blank)
    …next group…
    GRAND TOTAL

Design rules, each of which is load-bearing:

* Groups map to summary rows by matching the prefix against the row's Project
  Name; the un-prefixed group takes the remaining row. Nothing is keyed on a
  hard-coded PR code, because SoftServe reissues those every month.
* Ambiguity is fatal, never guessed. An unmapped prefix or a row-count mismatch
  aborts with the actual names printed, so a wrong project never gets billed.
* --bt is mandatory whenever the source workbook has a Business trip row.
  Carrying last month's travel cost forward silently would be a billing error,
  so the caller has to say keep or drop out loud.
* Dates are written as dd/mm/yyyy TEXT. Excel parses "07/06/2026" as 7 June in
  a US locale and silently mangles every day <= 12 (the June sheet has seven such
  corrupted cells). Text side-steps it and matches the sheet's own convention.
* Amounts are formulas, not baked numbers, so the sheet stays auditable:
  =RATE*ROUND(<hours cell>*24,2).
"""
import argparse
import datetime
import json
import re
import sys
import zipfile
from copy import copy

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

COLS = 'ABCDEF'
EXPECTED_HEADER = ['Description', 'Start Date', 'Start Time', 'End Time',
                   'Duration (h)', 'Amount, USD']
PREFIX_RE = re.compile(r'^\[([^\]]+)\]\s*')
DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}$')
TIME_FMT = 'h:mm:ss'                       # entry Start/End/Duration cells (C, D, E)
BT_SHEET = 'BT details'
R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'


def die(msg):
    sys.exit(f'report_build: {msg}')


def hhmmss(s):
    h, m, sec = map(int, s.split(':'))
    return datetime.time(h, m, sec)


# ---------------------------------------------------------------- style capture
def snapshot(ws, row):
    return {c.column_letter: (copy(c.font), copy(c.fill), copy(c.border),
                              copy(c.alignment), c.number_format) for c in ws[row]}


def stamp(cell, tpl, col):
    if col not in tpl:
        return
    f, fl, b, al, nf = tpl[col]
    cell.font, cell.fill, cell.border, cell.alignment = copy(f), copy(fl), copy(b), copy(al)
    cell.number_format = nf


def find_entry_row(ws):
    """Row number of the first genuine TIME-ENTRY row, or None.

    An entry row is identified by its own data, not its position: column B holds a
    dd/mm/yyyy TEXT date and column C holds a time. Never trust row 2 for this --
    in any workbook this script has already built, row 2 is the grey group band, and
    snapshotting it hands every entry the band's bold font, grey fill and 'General'
    format (which is exactly what happened to Aug'26; see the module docstring).
    """
    for r in range(2, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if isinstance(b, str) and DATE_RE.match(b.strip()) and c is not None:
            return r
    return None


def find_row(ws, *labels):
    """First row whose column A matches any label (case-insensitive, prefix-ok)."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str):
            low = v.strip().lower()
            for lab in labels:
                if low == lab.lower() or low.startswith(lab.lower()):
                    return r
    return None


# ---------------------------------------------------------------- group mapping
def group_entries(entries):
    """-> ordered dict {prefix or '' : [entries]}, preserving first-seen order."""
    groups = {}
    for e in entries:
        m = PREFIX_RE.match(e['desc'])
        groups.setdefault(m.group(1) if m else '', []).append(e)
    return groups


def map_groups_to_rows(groups, summary_rows):
    """prefix -> summary row. Match on project name; '' takes what is left over."""
    mapping, taken = {}, set()
    for prefix in [p for p in groups if p]:
        hits = [r for r in summary_rows
                if prefix.lower() in (r['name'] or '').lower() and r['row'] not in taken]
        if len(hits) != 1:
            die(f'cannot map prefix [{prefix}] to a project row -- '
                f'{len(hits)} name match(es) among '
                f'{[r["name"] for r in summary_rows]}. Fix the summary tab, or '
                f'rename the Clockify prefix, then re-run.')
        mapping[prefix] = hits[0]
        taken.add(hits[0]['row'])

    if '' in groups:
        left = [r for r in summary_rows if r['row'] not in taken]
        if len(left) != 1:
            die(f'the un-prefixed entries cannot be placed: {len(left)} summary row(s) '
                f'left unclaimed ({[r["name"] for r in left]}). Expected exactly 1.')
        mapping[''] = left[0]
        taken.add(left[0]['row'])

    unused = [r for r in summary_rows if r['row'] not in taken]
    if unused:
        die(f'summary tab has project row(s) with no matching time entries: '
            f'{[r["name"] for r in unused]}. Remove them or log time against them.')
    return mapping


# ---------------------------------------------------------------- OOXML salvage
def reinject(original, rebuilt):
    """openpyxl drops SharePoint customXml + printerSettings on round-trip; restore them.

    (calcChain and sharedStrings are also dropped and are deliberately NOT restored --
    Excel regenerates the first and openpyxl writes inline strings instead of the second.)
    """
    zo, zn = zipfile.ZipFile(original), zipfile.ZipFile(rebuilt)
    parts = {n: zn.read(n) for n in zn.namelist()}
    extra = [n for n in zo.namelist()
             if n.startswith('customXml/') or n.startswith('xl/printerSettings/')]
    if not extra:
        zo.close(); zn.close()
        return
    for n in extra:
        parts[n] = zo.read(n)

    rels = parts['xl/_rels/workbook.xml.rels'].decode()
    used = set(re.findall(r'Id="(rId\d+)"', rels))
    add = ''
    for n in sorted(x for x in extra if re.fullmatch(r'customXml/item\d+\.xml', x)):
        rid = next(f'rId{i}' for i in range(1, 999) if f'rId{i}' not in used)
        used.add(rid)
        add += f'<Relationship Id="{rid}" Type="{R_NS}/customXml" Target="../{n}"/>'
    parts['xl/_rels/workbook.xml.rels'] = rels.replace('</Relationships>',
                                                       add + '</Relationships>').encode()

    if any(x.startswith('xl/printerSettings/') for x in extra):
        s1 = parts['xl/worksheets/sheet1.xml'].decode()
        if '<pageSetup' in s1 and 'r:id=' not in s1:
            if 'xmlns:r=' not in s1:
                s1 = s1.replace(
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
                    f' xmlns:r="{R_NS}"', 1)
            s1 = re.sub(r'<pageSetup([^>]*?)\s*/>', r'<pageSetup\1 r:id="rId1"/>', s1, count=1)
            parts['xl/worksheets/sheet1.xml'] = s1.encode()
            parts['xl/worksheets/_rels/sheet1.xml.rels'] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                f'<Relationship Id="rId1" Type="{R_NS}/printerSettings" '
                'Target="../printerSettings/printerSettings1.bin"/></Relationships>').encode()

    ct = parts['[Content_Types].xml'].decode()
    ins = ''
    for n in sorted(x for x in extra if re.fullmatch(r'customXml/itemProps\d+\.xml', x)):
        ov = (f'<Override PartName="/{n}" ContentType="application/vnd.openxmlformats-'
              'officedocument.customXmlProperties+xml"/>')
        if ov not in ct:
            ins += ov
    if 'Extension="bin"' not in ct and any(x.endswith('.bin') for x in extra):
        ins += ('<Default Extension="bin" ContentType="application/vnd.openxmlformats-'
                'officedocument.spreadsheetml.printerSettings"/>')
    parts['[Content_Types].xml'] = ct.replace('</Types>', ins + '</Types>').encode()
    zo.close(); zn.close()

    order = ['[Content_Types].xml'] + [n for n in parts if n != '[Content_Types].xml']
    with zipfile.ZipFile(rebuilt, 'w', zipfile.ZIP_DEFLATED) as z:
        for n in order:
            z.writestr(n, parts[n])


# ---------------------------------------------------------------------- builder
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--entries', required=True)
    ap.add_argument('--template', required=True, help='workbook to clone structure/styles from')
    ap.add_argument('--out', required=True)
    ap.add_argument('--month', required=True, help='YYYY-MM being reported')
    ap.add_argument('--rate', type=float, default=120)
    ap.add_argument('--bt', choices=['keep', 'drop'],
                    help='what to do with the template Business trip row (required if present)')
    ap.add_argument('--bt-group', default='',
                    help="prefix of the group the trip belongs to ('' = un-prefixed group)")
    a = ap.parse_args()

    payload = json.load(open(a.entries))
    entries = payload['entries']
    if not entries:
        die('entries file is empty -- nothing to bill')

    # Guard against billing the wrong month: a stale export in ~/Downloads, or an API
    # fetch for the wrong period, otherwise produces a perfectly plausible workbook.
    want_y, want_m = (int(x) for x in a.month.split('-'))
    stray = sorted({e['date'] for e in entries
                    if datetime.datetime.strptime(e['date'], '%d/%m/%Y').year != want_y
                    or datetime.datetime.strptime(e['date'], '%d/%m/%Y').month != want_m})
    if stray:
        die(f'{len(stray)} entry date(s) fall outside {a.month}: {stray[:5]}'
            f'{"…" if len(stray) > 5 else ""}. Wrong timesheet for this month -- '
            f'check the source file.')

    wb = openpyxl.load_workbook(a.template)
    if 'Details' not in wb.sheetnames:
        die(f'template has no "Details" tab (tabs: {wb.sheetnames})')
    ws, summary = wb['Details'], wb[wb.sheetnames[0]]

    header = [ws.cell(row=1, column=i).value for i in range(1, 7)]
    if header != EXPECTED_HEADER:
        die(f'template Details header is {header}, expected {EXPECTED_HEADER}. '
            f"Use a Jun'26-or-later workbook as the template.")

    # --- summary project rows (source of truth for PR codes) --------------------
    rows = []
    for r in range(2, summary.max_row + 1):
        pr = summary.cell(row=r, column=4).value
        if isinstance(pr, str) and pr.strip():
            rows.append({'row': r, 'pr': pr.strip(),
                         'name': (summary.cell(row=r, column=5).value or '').strip()})
    if not rows:
        die('summary tab has no project rows (column D empty) -- '
            'add the PR codes SoftServe issued for this month first')

    groups = group_entries(entries)
    mapping = map_groups_to_rows(groups, rows)

    # --- business trip: preserved verbatim, never carried forward silently ------
    bt_row = find_row(ws, 'Business trip')
    bt_vals = None
    if bt_row:
        if not a.bt:
            die(f'template row {bt_row} has a Business trip of '
                f'{ws.cell(row=bt_row, column=6).value!r}. Pass --bt keep (it belongs to '
                f'{a.month}) or --bt drop (it was last month\'s). Refusing to guess.')
        if a.bt == 'keep':
            bt_vals = [ws.cell(row=bt_row, column=i).value for i in range(1, 7)]

    # --- capture styles before clearing ----------------------------------------
    entry_row = find_entry_row(ws)
    if entry_row is None:
        die('no time-entry row found in the template Details tab (looked for a '
            'dd/mm/yyyy date in column B with a time in C). Refusing to guess an '
            'entry style -- the sheet layout has changed.')
    tpl_entry = snapshot(ws, entry_row)
    tot_row = find_row(ws, 'Subtotal', 'Total', 'GRAND TOTAL')
    tpl_total = snapshot(ws, tot_row) if tot_row else snapshot(ws, 1)
    tpl_bt = snapshot(ws, bt_row) if bt_row else tpl_entry
    grand_row = find_row(ws, 'GRAND TOTAL')
    tpl_grand = snapshot(ws, grand_row) if grand_row else tpl_total

    ws.delete_rows(2, ws.max_row)
    grey = PatternFill('solid', fgColor='FFD9D9D9')
    thin = Border(*[Side(style='thin')] * 4)

    def band(r, text):
        for col in COLS:
            c = ws[f'{col}{r}']
            c.fill, c.border, c.font = grey, thin, Font(name='Arial', size=10, bold=True)
        ws[f'A{r}'] = text

    def totrow(r, label, e_formula, f_formula, tpl=None):
        ws[f'A{r}'] = label
        if e_formula:
            ws[f'E{r}'] = e_formula
        ws[f'F{r}'] = f_formula
        for col in COLS:
            stamp(ws[f'{col}{r}'], tpl or tpl_total, col)
        ws[f'E{r}'].number_format = '[h]:mm:ss'

    r = 2
    hour_cells, group_cells = [], []
    for target in rows:                                  # summary order drives sheet order
        prefix = next(p for p, m in mapping.items() if m['row'] == target['row'])
        band(r, f"{target['name']}  ({target['pr']})")
        r += 1
        first = r
        for e in groups[prefix]:
            ws.cell(row=r, column=1, value=e['desc'])
            ws.cell(row=r, column=2, value=e['date'])     # TEXT -- see module docstring
            ws.cell(row=r, column=3, value=hhmmss(e['start']))
            ws.cell(row=r, column=4, value=hhmmss(e['end']))
            ws.cell(row=r, column=5, value=hhmmss(e['dur']))
            for col in COLS:
                stamp(ws[f'{col}{r}'], tpl_entry, col)
            # Enforce the time format explicitly. openpyxl assigns it when a
            # datetime.time is written, but stamp() then overwrites number_format
            # from the template row -- so a template whose entry style ever drifts
            # to 'General' would silently render these as 0.08333333 decimals.
            for col in 'CDE':
                ws[f'{col}{r}'].number_format = TIME_FMT
            r += 1
        last = r - 1

        has_bt = bt_vals is not None and prefix == a.bt_group
        hours_r = r
        totrow(hours_r, 'Subtotal - hours' if has_bt else f"Subtotal - {target['pr']}",
               f'=SUM(E{first}:E{last})', f'={a.rate:g}*ROUND(E{hours_r}*24,2)')
        r += 1
        hour_cells.append(f'E{hours_r}')
        money_cell = f'F{hours_r}'

        if has_bt:
            for i, v in enumerate(bt_vals, start=1):
                ws.cell(row=r, column=i, value=v)
            for col in COLS:
                stamp(ws[f'{col}{r}'], tpl_bt, col)
            bt_at = r
            r += 1
            totrow(r, f"Subtotal - {target['pr']}", f'=E{hours_r}', f'=F{hours_r}+F{bt_at}')
            money_cell = f'F{r}'
            r += 1

        group_cells.append(money_cell)
        summary.cell(row=target['row'], column=6).value = f'=Details!{money_cell}'
        r += 1                                            # blank spacer after the group

    grand = r                                             # ...so GRAND TOTAL sits below it
    ws[f'A{grand}'] = 'GRAND TOTAL'
    ws[f'E{grand}'] = '=' + '+'.join(hour_cells)
    ws[f'F{grand}'] = '=' + '+'.join(group_cells)
    for col in COLS:
        stamp(ws[f'{col}{grand}'], tpl_grand, col)
    ws[f'E{grand}'].number_format = '[h]:mm:ss'

    # --- summary period ---------------------------------------------------------
    y, mo = (int(x) for x in a.month.split('-'))
    for target in rows:
        summary.cell(row=target['row'], column=1).value = datetime.datetime(y, mo, 1)

    # --- BT details: clear it when there is no trip this month -------------------
    # 'BT details' is the supporting maths for the Business trip line. With --bt keep
    # it is Alex's hand-built sheet and must never be touched. With --bt drop there is
    # no trip, so leaving the template's numbers in place ships LAST month's travel
    # cost to the client as apparent support for a line that no longer exists --
    # the same billing error --bt exists to prevent (Aug'26 shipped once with July's
    # "Stay at Wroclaw" maths intact). Clear the figures, keep the scaffolding:
    # row 1 headers stay, and plain text labels from column D on stay, so the sheet
    # stays a ready template for the next trip.
    bt_cleared = 0
    if a.bt == 'drop' and BT_SHEET in wb.sheetnames:
        bts = wb[BT_SHEET]
        for row in bts.iter_rows(min_row=2, max_row=bts.max_row):
            for c in row:
                if c.value is None:
                    continue
                is_label = (isinstance(c.value, str) and not c.value.startswith('=')
                            and c.column >= 4)
                if not is_label:
                    c.value = None
                    bt_cleared += 1

    wb.save(a.out)
    reinject(a.template, a.out)

    print(json.dumps({
        'out': a.out, 'month': a.month, 'rate': a.rate,
        'entries': len(entries), 'period_total': payload.get('period_total'),
        'business_trip': (bt_vals[5] if bt_vals else None),
        'bt_details_cells_cleared': bt_cleared,
        'groups': [{'prefix': p or '(none)', 'pr': mapping[p]['pr'],
                    'name': mapping[p]['name'], 'entries': len(g)}
                   for p, g in groups.items()],
        'grand_total_row': grand,
    }, indent=1, ensure_ascii=False))


if __name__ == '__main__':
    main()
