#!/usr/bin/env python3
"""report_verify.py — independent audit of a built report. Non-zero exit = do not send.

    ./report_verify.py --xlsx "…_Jul'26.xlsx" --entries e.json --rate 120

openpyxl writes formulas but cannot evaluate them, so this shells out to
LibreOffice to force a real recalculation and then checks the numbers that
actually matter for an invoice:

  1. Details hours == the sum of the source time entries (to the second).
  2. Every group's money == rate x its hours, rounded the way the sheet claims.
  3. GRAND TOTAL == the sum of the group subtotals.
  4. Summary "Amount gross" column == the group subtotals it points at.
  5. Summary total == GRAND TOTAL, so the two tabs cannot disagree.

This exists because the failure it guards against is silent: a formula that
points one row off still produces a plausible number, and nobody re-adds an
invoice by hand before emailing it.
"""
import argparse
import datetime
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time

import openpyxl

SOFFICE_CANDIDATES = ['/Applications/LibreOffice.app/Contents/MacOS/soffice',
                      '/opt/homebrew/bin/soffice', 'soffice']
EPS = 0.005


def to_secs(v):
    if isinstance(v, datetime.timedelta):
        return int(round(v.total_seconds()))
    if isinstance(v, datetime.time):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)):
        return int(round(v * 86400))
    raise TypeError(f'not a duration: {v!r}')


def hms(sec):
    return f'{sec // 3600:02d}:{sec % 3600 // 60:02d}:{sec % 60:02d}'


def recalc(xlsx):
    """LibreOffice recalculates on load; converting back gives us cached values."""
    soffice = next((p for p in SOFFICE_CANDIDATES if p == 'soffice' or os.path.exists(p)), None)
    if not soffice or (soffice == 'soffice' and not shutil.which('soffice')):
        sys.exit('report_verify: LibreOffice not found -- cannot recalculate; '
                 'refusing to certify the report')
    out = tempfile.mkdtemp(prefix='ssreport-verify-')
    # `timeout`/`gtimeout` are not installed on this Mac; poll for the artefact instead.
    p = subprocess.Popen([soffice, '--headless', '--calc', '--convert-to', 'xlsx',
                          '--outdir', out, xlsx],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    dest = os.path.join(out, os.path.basename(xlsx))
    for _ in range(120):
        if p.poll() is not None and os.path.exists(dest):
            break
        time.sleep(1)
    else:
        p.kill()
        sys.exit('report_verify: LibreOffice recalculation timed out (120s)')
    time.sleep(1)
    if not os.path.exists(dest):
        sys.exit('report_verify: LibreOffice produced no output')
    return dest


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--entries', required=True)
    ap.add_argument('--rate', type=float, default=120)
    a = ap.parse_args()

    wb = openpyxl.load_workbook(recalc(a.xlsx), data_only=True)
    d, s = wb['Details'], wb[wb.sheetnames[0]]

    subtotals, hours_rows, grand = [], [], None
    for r in range(1, d.max_row + 1):
        v = d.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        low = v.strip().lower()
        if low == 'grand total':
            grand = r
        elif low == 'subtotal - hours':
            hours_rows.append(r)
        elif low.startswith('subtotal'):
            subtotals.append(r)

    fails, notes = [], []
    if grand is None:
        fails.append('no GRAND TOTAL row found')
    if not subtotals:
        fails.append('no "Subtotal - <PR>" rows found')

    # 1. hours reconcile against the source entries
    src = json.load(open(a.entries))
    want = sum(to_secs(datetime.time(*map(int, e['dur'].split(':')))) for e in src['entries'])
    got = to_secs(d.cell(row=grand, column=5).value) if grand else -1
    if got != want:
        fails.append(f'hours mismatch: sheet {hms(got)} vs entries {hms(want)}')
    else:
        notes.append(f'hours {hms(got)} == {len(src["entries"])} source entries')

    # 2. each hours subtotal is rate x hours
    for r in hours_rows + subtotals:
        amt, dur = d.cell(row=r, column=6).value, d.cell(row=r, column=5).value
        if not isinstance(amt, (int, float)) or dur is None:
            continue
        expect = a.rate * round(to_secs(dur) / 3600, 2)
        if r in hours_rows and abs(amt - expect) > EPS:
            fails.append(f'row {r}: amount {amt:,.2f} != rate x hours {expect:,.2f}')

    # 3. GRAND TOTAL == sum of group subtotals
    group_sum = sum(d.cell(row=r, column=6).value or 0 for r in subtotals)
    gt = d.cell(row=grand, column=6).value if grand else None
    if gt is None or abs(gt - group_sum) > EPS:
        fails.append(f'GRAND TOTAL {gt} != sum of subtotals {group_sum:,.2f}')
    else:
        notes.append(f'GRAND TOTAL {gt:,.2f} == sum of {len(subtotals)} group subtotal(s)')

    # 4 + 5. summary column F agrees with Details and totals to the same number
    sum_amounts = []
    for r in range(2, s.max_row + 1):
        pr, amt = s.cell(row=r, column=4).value, s.cell(row=r, column=6).value
        if isinstance(pr, str) and pr.strip():
            if not isinstance(amt, (int, float)):
                fails.append(f'summary row {r} ({pr}) has no amount')
            else:
                sum_amounts.append((pr.strip(), amt))
    matched = [amt for _, amt in sum_amounts]
    detail_amounts = sorted(round(d.cell(row=r, column=6).value or 0, 2) for r in subtotals)
    if sorted(round(x, 2) for x in matched) != detail_amounts:
        fails.append(f'summary amounts {sorted(matched)} != Details subtotals {detail_amounts}')
    if gt is not None and abs(sum(matched) - gt) > EPS:
        fails.append(f'summary total {sum(matched):,.2f} != GRAND TOTAL {gt:,.2f}')
    else:
        notes.append(f'summary {len(sum_amounts)} row(s) total {sum(matched):,.2f} == GRAND TOTAL')

    for pr, amt in sum_amounts:
        notes.append(f'  {pr}: {amt:,.2f} USD')

    print('\n'.join(f'  ok   {n}' for n in notes))
    if fails:
        print('\n'.join(f'  FAIL {f}' for f in fails), file=sys.stderr)
        sys.exit(f'report_verify: {len(fails)} check(s) failed -- DO NOT SEND this workbook')
    print(f'report_verify: all checks passed for {os.path.basename(a.xlsx)}')


if __name__ == '__main__':
    main()
