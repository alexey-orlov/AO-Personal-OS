# -*- coding: utf-8 -*-
"""Build the final ranked workbook (xlsx with tabs) + a CSV for the repo."""
import csv, json, math, os, sys

HERE = "/tmp/claude-0/-home-user-AO-Personal-OS/d13b3c34-2805-51c9-80da-5179e1911a53/scratchpad"
sys.path.insert(0, HERE)
from score import build

REPO = "/home/user/AO-Personal-OS/context/areas/job-search/docs/bay-area-relocation"

SIGMA = 0.28  # log-sd of asking rents within a market/bedroom class


def share_under(median, threshold=4000.0, sigma=SIGMA):
    """P(rent < threshold) under a log-normal centred on the reported median/average."""
    if not median or median <= 0:
        return None
    z = (math.log(threshold) - math.log(median)) / sigma
    return 0.5 * (1 + math.erf(z / math.sqrt(2)))


def band(share):
    if share is None:
        return "-"
    if share >= 0.60:
        return "Plenty"
    if share >= 0.30:
        return "Some"
    if share >= 0.10:
        return "Few"
    if share >= 0.02:
        return "Very few"
    return "Almost none"


COLUMNS = [
    ("rank", "Rank"),
    ("place", "Place"),
    ("county", "County"),
    ("kind_label", "Type"),
    ("total_r", "TOTAL SCORE"),
    ("c_safety_r", "1. Safety"),
    ("c_schools_r", "2. Schools"),
    ("c_rent_r", "3. Rent cost"),
    ("c_distance_r", "4. Close to SF"),
    ("c_capacity_r", "5. School/TK capacity"),
    ("c_urban_r", "6. Urban"),
    ("rent_2br_r", "2BR median rent $"),
    ("rent_3br_r", "3BR median rent $"),
    ("lst2_total_r", "2BR listings (live count)"),
    ("u4k_2_count", "2BR under $4k (est. count)"),
    ("u4k_2_band", "2BR under $4k supply"),
    ("lst3_total_r", "3BR listings (live count)"),
    ("u4k_3_count", "3BR under $4k (est. count)"),
    ("u4k_3_band", "3BR under $4k supply"),
    ("drive_min", "Drive to SF (min, off-peak)"),
    ("gc_mi", "Miles to SF (straight line)"),
    ("violent_r", "Violent crime /1k"),
    ("property_r", "Property crime /1k"),
    ("elem_district", "School district"),
    ("elem_rating_r", "Elementary rating /10"),
    ("high_rating_r", "Middle+High rating /10"),
    ("assignment", "Assignment system"),
    ("enroll_trend", "Enrolment trend"),
    ("capacity_pressure", "Seat availability"),
    ("tk_note", "TK capacity"),
    ("pop", "Population"),
    ("completeness", "Data completeness %"),
    ("imputed_s", "Imputed cells"),
    ("sources_s", "Sources"),
    ("notes_s", "Notes"),
]

KIND_LABEL = {"sf_hood": "SF neighbourhood", "city": "City", "town": "Town",
              "cdp": "Unincorporated community"}


def r1(x):
    return round(x, 1) if isinstance(x, (int, float)) else x


def prep(rows):
    out = []
    for r in rows:
        d = dict(r)
        d["kind_label"] = KIND_LABEL.get(r["kind"], r["kind"])
        for k in ["total", "c_safety", "c_schools", "c_rent", "c_distance", "c_capacity",
                  "c_urban", "violent", "property", "elem_rating", "high_rating"]:
            d[k + "_r"] = r1(r.get(k))
        for k in ["rent_2br", "rent_3br", "lst2_total", "lst3_total"]:
            v = r.get(k)
            d[k + "_r"] = round(v) if isinstance(v, (int, float)) else v

        s2 = share_under(r.get("rent_2br"))
        s3 = share_under(r.get("rent_3br"))
        d["u4k_2_band"] = band(s2)
        d["u4k_3_band"] = band(s3)
        d["u4k_2_count"] = (round(r["lst2_total"] * s2) if r.get("lst2_total") and s2 is not None else "-")
        d["u4k_3_count"] = (round(r["lst3_total"] * s3) if r.get("lst3_total") and s3 is not None else "-")

        srcs = []
        for k in ["rent_2br_src", "rent_3br_src", "lst2_src", "lst3_src", "crime_src"]:
            v = r.get(k)
            if v and v not in srcs:
                srcs.append(v)
        d["sources_s"] = "; ".join(srcs) if srcs else "-"
        d["imputed_s"] = ", ".join(r.get("imputed", [])) or "-"
        n = (r.get("notes") or "").strip()
        q = (r.get("crime_qual") or "").strip()
        both = " | ".join(x for x in [n, q] if x)
        d["notes_s"] = both[:900] if both else "-"
        for k, _ in COLUMNS:
            if d.get(k) is None:
                d[k] = "-"
        out.append(d)
    return out


def main():
    rows, live = build()
    nonres = [r for r in rows if not r["residential"]]
    plive, pnon = prep(live), prep(nonres)

    # ---- CSV (full detail) for the repo
    os.makedirs(REPO, exist_ok=True)
    csv_path = os.path.join(REPO, "bay-area-ranking.csv")
    with open(csv_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow([h for _, h in COLUMNS])
        for d in plive + pnon:
            w.writerow([d.get(k, "-") for k, _ in COLUMNS])
    print("wrote", csv_path, len(plive) + len(pnon), "rows")

    # ---- xlsx with tabs
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"
    ws.append([h for _, h in COLUMNS])
    for d in plive:
        ws.append([d.get(k, "-") for k, _ in COLUMNS])
    for d in pnon:
        ws.append([d.get(k, "-") for k, _ in COLUMNS])

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    widths = {"Place": 30, "County": 13, "Type": 22, "School district": 34,
              "Assignment system": 26, "Enrolment trend": 18, "Seat availability": 18,
              "TK capacity": 30, "Sources": 30, "Notes": 70, "Imputed cells": 22}
    for i, (_, h) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 12)
    ws.row_dimensions[1].height = 58

    # score colour scale on the total
    from openpyxl.formatting.rule import ColorScaleRule
    tot_col = get_column_letter([h for _, h in COLUMNS].index("TOTAL SCORE") + 1)
    ws.conditional_formatting.add(
        f"{tot_col}2:{tot_col}{len(plive)+1}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))
    for cl in ["F", "G", "H", "I", "J", "K"]:
        ws.conditional_formatting.add(
            f"{cl}2:{cl}{len(plive)+1}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))

    # ---- method tab
    m = wb.create_sheet("How to read this")
    for line in METHOD_ROWS:
        m.append(line)
    m.column_dimensions["A"].width = 30
    m.column_dimensions["B"].width = 120
    for c in m["A"]:
        c.font = Font(bold=True)
    for row in m.iter_rows(min_col=2, max_col=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    xlsx_path = os.path.join(HERE, "bay-area-ranking.xlsx")
    wb.save(xlsx_path)
    print("wrote", xlsx_path, os.path.getsize(xlsx_path), "bytes")
    return xlsx_path, csv_path, plive, pnon


METHOD_ROWS = [
    ["Bay Area relocation ranking", "220 places: 41 SF neighbourhoods + all 100 other incorporated Bay Area cities/towns + 79 unincorporated communities. Built 2026-08-09."],
    ["", ""],
    ["Scoring", "Six criteria, each weighted equally at 1/6. Each is converted to a 0-100 percentile rank within the dataset before averaging, so no single criterion's long tail can dominate. Higher is better in every column."],
    ["1. Safety", "Violent crime rate per 1,000 residents (65% of the criterion) + property crime rate per 1,000 (35%), inverted. Sources: police/city reports where available, otherwise CrimeGrade / NeighborhoodScout / AreaVibes. For SF neighbourhoods the aggregators are the only source published at neighbourhood resolution."],
    ["2. Schools", "Average of an elementary rating and a middle/high rating (1-10), because the household has one TK/K-age child and one middle/high-age child. Captured at DISTRICT level, not per school -- inside a big district (SFUSD, Oakland, San Jose) individual schools vary far more than districts do."],
    ["3. Rent cost", "Median 2BR and 3BR rent, averaged, inverted so cheaper scores higher."],
    ["4. Close to SF", "Modelled off-peak drive minutes to the Ferry Building. No routing service was reachable, so time is fitted per drive corridor as t = a + b x straight-line distance, calibrated on 81 anchor trips (mean absolute error 1.2 min). Rush hour is NOT modelled and diverges a lot by corridor."],
    ["5. School/TK capacity", "Ease of actually getting a free public-school seat and a TK seat: district enrolment trend (declining = open seats), assignment system (strict address-based boundary = +8, choice lottery = -18), and whether TK space is constrained (-12) or widely available (+6)."],
    ["6. Urban", "Distance-decayed population potential -- population of every other place weighted by exp(-distance/8km), with each place's population spread over a footprint rather than sitting at a point. More urban scores higher, per the brief."],
    ["", ""],
    ["Sub-$4,000 supply", "The 'live count' columns are real listing counts read from rental-site pages. The 'under $4k (est. count)' columns are MODELLED: price-filtered pages are not indexed by search, so the share of inventory below $4,000 is estimated from the market's median rent using a log-normal price distribution (sigma 0.28) and multiplied by the live count. Trust the supply BAND more than the point estimate."],
    ["Data completeness %", "Share of the six measured inputs (2BR rent, 3BR rent, violent crime, property crime, elementary rating, middle/high rating) that were actually measured rather than imputed. Rows below ~50% are reconstructions -- almost all are small unincorporated communities with no rental market or crime reporting of their own."],
    ["Imputed cells", "Names the fields filled from the county+place-type median because no source could be found. Never silently filled."],
    ["", ""],
    ["Collection limits", "This ran in an environment whose network policy blocked direct HTTP and the URL-fetching tool for nearly every domain, so Census, HUD, California DOJ, CDE and DataSF bulk files were unavailable. Everything researched came via web search, with the publisher domain stored per figure."],
    ["Not covered", "Buying (rent only), public transit, climate/wildfire/flood risk, air quality, commute to any employer other than downtown SF, and the character of a place beyond its density."],
    ["Equal weighting", "Equal weighting is the brief, not a recommendation. Every component column is kept so the weights can be changed."],
    ["Non-residential rows", "Golden Gate Park, Lincoln Park, McLaren Park and the Presidio are parkland with negligible housing. They are listed at the bottom without a rank rather than given a misleading score."],
]

if __name__ == "__main__":
    main()
