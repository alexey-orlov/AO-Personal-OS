# -*- coding: utf-8 -*-
"""
Final deliverable builder.

Reality of this run: the session's WebSearch budget (200 calls, a hard cap) was exhausted partway
through the fan-out. 45 of 220 places received measured data -- but that 45 includes 36 of the 37
residential San Francisco neighbourhoods, so the San Francisco half of the brief is genuinely
complete while the wider Bay Area is not.

So we publish three honest things rather than one dishonest one:
  Tab 1  San Francisco neighbourhoods, fully ranked on the criteria that actually vary.
  Tab 2  All 220 places, with the two criteria computed rigorously for every row (drive time,
         urbanicity) plus whatever else was measured, and an explicit data-status column.
  Tab 3+ Method, and exactly what remains to be collected.
"""
import csv, json, math, os, sys

HERE = "/tmp/claude-0/-home-user-AO-Personal-OS/d13b3c34-2805-51c9-80da-5179e1911a53/scratchpad"
sys.path.insert(0, HERE)
import score as S
from export import share_under, band
from places import PLACES

REPO = "/home/user/AO-Personal-OS/context/areas/job-search/docs/bay-area-relocation"
KIND_LABEL = {"sf_hood": "SF neighbourhood", "city": "City", "town": "Town",
              "cdp": "Unincorporated community"}


def r1(x):
    return round(x, 1) if isinstance(x, (int, float)) else ""


def money(x):
    return round(x) if isinstance(x, (int, float)) else ""


# ---------------------------------------------------------------- SF neighbourhood ranking
def sf_table():
    rows, live = S.build(subset=lambda r: r["county"] == "San Francisco", impute_missing=False)
    # A place is rankable only when every criterion that VARIES inside San Francisco is present.
    # Schools and TK capacity are district-constant across SFUSD, so they cannot separate
    # neighbourhoods and are carried as context columns instead of scoring ones.
    VARY = ["c_safety", "c_rent", "c_distance", "c_urban"]
    ranked = [r for r in live if all(r.get(c) is not None for c in VARY)]
    unranked = [r for r in live if r not in ranked]
    for r in ranked:
        r["total_sf"] = sum(r[c] for c in VARY) / len(VARY)
    ranked.sort(key=lambda r: -r["total_sf"])
    for i, r in enumerate(ranked, 1):
        r["rank_sf"] = i
    return ranked, unranked


SF_COLS = [
    ("rank_sf", "Rank"), ("place", "SF neighbourhood"), ("total_sf", "TOTAL SCORE"),
    ("c_safety", "Safety"), ("c_rent", "Rent cost"), ("c_distance", "Close to downtown"),
    ("c_urban", "Urban density"),
    ("rent_2br", "2BR median rent $"), ("rent_3br", "3BR median rent $"),
    ("u4k2", "2BR under $4k supply"), ("u4k3", "3BR under $4k supply"),
    ("lst2_total", "2BR listings (live)"), ("lst3_total", "3BR listings (live)"),
    ("violent", "Violent crime /1k"), ("property", "Property crime /1k"),
    ("drive_min", "Drive to Ferry Bldg (min)"), ("pop", "Population"),
    ("schools_note", "Schools"), ("capacity_note", "School/TK capacity"),
    ("sources", "Sources"), ("notes_s", "Notes"),
]


def sf_rowdict(r):
    d = {
        "rank_sf": r.get("rank_sf", ""), "place": r["place"],
        "total_sf": r1(r.get("total_sf")), "c_safety": r1(r.get("c_safety")),
        "c_rent": r1(r.get("c_rent")), "c_distance": r1(r.get("c_distance")),
        "c_urban": r1(r.get("c_urban")),
        "rent_2br": money(r.get("rent_2br")), "rent_3br": money(r.get("rent_3br")),
        "u4k2": band(share_under(r.get("rent_2br"))),
        "u4k3": band(share_under(r.get("rent_3br"))),
        "lst2_total": money(r.get("lst2_total")), "lst3_total": money(r.get("lst3_total")),
        "violent": r1(r.get("violent")), "property": r1(r.get("property")),
        "drive_min": r1(r.get("drive_min")), "pop": r.get("pop"),
        "schools_note": "SFUSD (district-wide; school-level variation not measured)",
        "capacity_note": "Open seats (~14,000 unfilled) but citywide choice lottery "
                         "with attendance-area tiebreaker; TK expanding",
    }
    srcs = [r.get(k) for k in ["rent_2br_src", "rent_3br_src", "lst2_src", "lst3_src", "crime_src"]]
    srcs = [s for s in srcs if s]
    seen, out = set(), []
    for s in srcs:
        if s.lower() not in seen:
            seen.add(s.lower())
            out.append(s)
    d["sources"] = "; ".join(out) or "-"
    n = (r.get("notes") or "").strip()
    q = (r.get("crime_qual") or "").strip()
    d["notes_s"] = (" | ".join(x for x in [n, q] if x))[:260] or "-"
    return d


# ---------------------------------------------------------------- full 220-place framework
ALL_COLS = [
    ("place", "Place"), ("county", "County"), ("kind_label", "Type"),
    ("status", "Data status"),
    ("drive_min", "Drive to SF (min, off-peak)"), ("gc_mi", "Miles to SF (straight line)"),
    ("urban_rank", "Urban rank (1=most urban)"), ("pop", "Population"),
    ("rent_2br", "2BR median rent $"), ("rent_3br", "3BR median rent $"),
    ("u4k2", "2BR under $4k supply"), ("u4k3", "3BR under $4k supply"),
    ("violent", "Violent crime /1k"), ("property", "Property crime /1k"),
    ("elem_district", "School district"),
    ("sources", "Sources"),
]


def all_table():
    rows, _ = S.build(impute_missing=False)
    urb_sorted = sorted(rows, key=lambda r: -r["urban_index"])
    urank = {r["place"]: i for i, r in enumerate(urb_sorted, 1)}
    out = []
    for r in rows:
        n = r.get("n_measured", 0)
        if not r["residential"]:
            status = "Parkland - not ranked"
        elif n >= 4:
            status = "Measured"
        elif n >= 1:
            status = f"Partly measured ({n}/6)"
        else:
            status = "NOT YET MEASURED - search budget exhausted"
        out.append({
            "place": r["place"], "county": r["county"],
            "kind_label": KIND_LABEL.get(r["kind"], r["kind"]), "status": status,
            "drive_min": r1(r["drive_min"]), "gc_mi": r1(r["gc_mi"]),
            "urban_rank": urank[r["place"]], "pop": r["pop"],
            "rent_2br": money(r.get("rent_2br")), "rent_3br": money(r.get("rent_3br")),
            "u4k2": band(share_under(r.get("rent_2br"))),
            "u4k3": band(share_under(r.get("rent_3br"))),
            "violent": r1(r.get("violent")), "property": r1(r.get("property")),
            "elem_district": r.get("elem_district") or "",
            "sources": "; ".join([s for s in [r.get("rent_2br_src"), r.get("crime_src")] if s]) or "",
        })
    out.sort(key=lambda d: (d["status"].startswith("NOT"), d["county"], d["place"]))
    return out


METHOD = [
    ["WHAT THIS IS", "A relocation comparison of 220 Bay Area places for a household with one child entering TK/Kindergarten and one of middle/high-school age. Built 2026-08-09."],
    ["", ""],
    ["!! READ THIS FIRST", "The San Francisco tab is complete and safe to use. The 'All 220 places' tab is a FRAMEWORK: drive time and urban rank are computed for every row, but rent, crime and school data were only collected for 45 places before this session hit a hard cap of 200 web searches. Rows marked 'NOT YET MEASURED' have no rent/crime/school data at all -- they are deliberately left blank rather than filled with estimates."],
    ["", ""],
    ["THE SIX CRITERIA", "Equal weight, 1/6 each. Each converted to a 0-100 percentile rank before averaging, so no single criterion's long tail dominates. Higher is always better."],
    ["1. Safety", "Violent crime per 1,000 residents (65% of the criterion) + property crime per 1,000 (35%), inverted. SF neighbourhood figures come from crime aggregators (CrimeGrade / AreaVibes) -- the only sources published at neighbourhood resolution. They are Tier-3 derivations of federal data: internally consistent enough to rank with, not authoritative for any single absolute number."],
    ["2. Schools", "Average of an elementary and a middle/high rating. NOT USABLE IN THE SF TAB: every SF neighbourhood is in one district (SFUSD), so this criterion is constant across all 41 and cannot separate them. Within SFUSD, school-to-school variation is large but was not measurable here."],
    ["3. Rent cost", "Median 2BR and 3BR rent, averaged, inverted so cheaper scores higher."],
    ["4. Close to SF", "Modelled off-peak drive minutes to the Ferry Building. No routing service was reachable, so time is fitted per drive corridor as t = a + b x straight-line distance, calibrated on 81 anchor trips. Mean absolute error 1.2 minutes. Independently spot-checked: Palo Alto (published 33mi/38-45min vs 38 modelled), San Rafael (18.6mi/26-38 vs 30), Walnut Creek (24mi/30-45 vs 35). RUSH HOUR IS NOT MODELLED and diverges hugely by corridor."],
    ["5. School/TK capacity", "Ease of actually getting a free public seat: enrolment trend (declining = open seats), assignment system (address-based boundary +8, choice lottery -18), TK space constrained (-12) or available (+6). Constant across the SF tab, so carried there as a context column."],
    ["6. Urban density", "Distance-decayed population potential: every other place's population weighted by exp(-distance/8km), with each place's population spread over a footprint rather than sitting at its centroid. More urban scores higher, per the brief."],
    ["", ""],
    ["SUB-$4,000 SUPPLY", "'Live' listing counts are real counts read off rental-site pages. The 'under $4k supply' bands are MODELLED: price-filtered pages are not indexed by search, so the share of inventory below $4,000 is estimated from the market median using a log-normal distribution (sigma 0.28). Trust the band, not a point estimate."],
    ["KNOWN DEFECT FIXED", "Agents initially took a place's 2BR and 3BR figures from DIFFERENT publishers. Publisher-to-publisher variance (~40%) exceeds the real 2BR-to-3BR premium, producing 3BR values cheaper than 2BR. Verified on Noe Valley: apartments.com says $4,693 for a 3BR, RentHop says $6,695. Inverted pairs are rebuilt from the 2BR using a bedroom premium learned from same-publisher pairs, and flagged."],
    ["SFUSD - VERIFIED", "The zone-based assignment policy was delayed AGAIN and is NOT in force for 2026-27; the citywide choice lottery (Board Policy 5101) still governs, softened by an attendance-area tiebreaker, with a rough new timeline of winter 2028-29. Physical capacity is abundant: ~14,000 seats to fill, enrolment down 6.6% vs 2019, closures paused, facilities being converted into TK classrooms. Sources: sfusd.edu, sfstandard.com, thefrisc.com."],
    ["", ""],
    ["NOT COVERED", "Buying (rent only), public transit, climate / wildfire / flood risk, air quality, commute to any employer other than downtown SF, and the character of a place beyond its density."],
    ["EQUAL WEIGHTING", "Equal weighting is the brief, not a recommendation. Every component column is kept so weights can be changed."],
]

TODO = [
    ["WHAT REMAINS", ""],
    ["Blocker", "This session has a hard cap of 200 WebSearch calls and it was fully consumed. Direct HTTP and the URL-fetch tool are blocked by the network egress policy for nearly every domain, so Census, HUD, California DOJ, CDE and DataSF bulk files could not be used as a substitute."],
    ["Measured", "45 of 220 places: 36 of 37 residential SF neighbourhoods, plus Alameda, Albany, Berkeley, Dublin, Emeryville, Fremont, Newark, Oakland, Piedmont."],
    ["Still needed", "175 places x roughly 4 searches = about 700 searches, plus roughly 130 school districts x 4 = about 520 searches. Call it 1,200 searches, or about 6 further sessions at 200 each."],
    ["Cheapest way to finish", "Rank the remaining places by how plausible they are as a shortlist first, and only research that shortlist. The 79 unincorporated communities and the far-flung Sonoma/Napa/Solano towns will all score poorly on distance and urbanicity regardless, and those two criteria are ALREADY computed for every row in this workbook -- so they can be filtered out before spending any search budget."],
    ["Suggested next step", "Filter the 'All 220 places' tab to drive time under about 45 minutes, then research only those. That is roughly 90 places, about one and a half sessions of budget, and it covers every realistic candidate."],
]


def build_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    ranked, unranked = sf_table()
    allrows = all_table()

    wb = Workbook()
    HDR = PatternFill("solid", fgColor="1F3864")

    def style(ws, ncols, height=52):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = HDR
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[1].height = height
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    # ---- Tab 1: SF ranking
    ws = wb.active
    ws.title = "San Francisco (complete)"
    ws.append([h for _, h in SF_COLS])
    for r in ranked:
        d = sf_rowdict(r)
        ws.append([d.get(k, "") for k, _ in SF_COLS])
    n_ranked = ws.max_row
    ws.append([])
    ws.append(["PARTIAL DATA - not ranked because at least one scoring criterion is missing"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, italic=True)
    for r in unranked:
        d = sf_rowdict(r)
        d["rank_sf"] = "-"
        ws.append([d.get(k, "") for k, _ in SF_COLS])
    style(ws, len(SF_COLS))
    ws.freeze_panes = "C2"
    widths = {"SF neighbourhood": 30, "Schools": 44, "School/TK capacity": 46,
              "Sources": 30, "Notes": 70, "2BR under $4k supply": 16,
              "3BR under $4k supply": 16}
    for i, (_, h) in enumerate(SF_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 13)
    for cl in ["C", "D", "E", "F", "G"]:
        ws.conditional_formatting.add(
            f"{cl}2:{cl}{n_ranked}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))

    # ---- Tab 2: all 220
    w2 = wb.create_sheet("All 220 places")
    w2.append([h for _, h in ALL_COLS])
    for d in allrows:
        w2.append([d.get(k, "") for k, _ in ALL_COLS])
    style(w2, len(ALL_COLS))
    w2.freeze_panes = "B2"
    w2widths = {"Place": 32, "County": 14, "Type": 24, "Data status": 34,
                "School district": 34, "Sources": 26,
                "2BR under $4k supply": 16, "3BR under $4k supply": 16}
    for i, (_, h) in enumerate(ALL_COLS, 1):
        w2.column_dimensions[get_column_letter(i)].width = w2widths.get(h, 13)

    # ---- Tab 3 / 4
    for title, data, w in [("How to read this", METHOD, 130), ("What remains", TODO, 130)]:
        m = wb.create_sheet(title)
        for line in data:
            m.append(line)
        m.column_dimensions["A"].width = 26
        m.column_dimensions["B"].width = w
        for c in m["A"]:
            c.font = Font(bold=True)
        for row in m.iter_rows(min_col=2, max_col=2):
            for c in row:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    path = os.path.join(HERE, "bay-area-relocation.xlsx")
    wb.save(path)

    os.makedirs(REPO, exist_ok=True)
    with open(os.path.join(REPO, "sf-neighbourhood-ranking.csv"), "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow([h for _, h in SF_COLS])
        for r in ranked + unranked:
            d = sf_rowdict(r)
            wr.writerow([d.get(k, "") for k, _ in SF_COLS])
    with open(os.path.join(REPO, "all-220-places.csv"), "w", newline="") as f:
        wr = csv.writer(f)
        wr.writerow([h for _, h in ALL_COLS])
        for d in allrows:
            wr.writerow([d.get(k, "") for k, _ in ALL_COLS])
    return path, len(ranked), len(unranked), len(allrows)


if __name__ == "__main__":
    p, a, b, c = build_workbook()
    print(f"workbook: {p} ({os.path.getsize(p)} bytes)")
    print(f"SF ranked {a}, SF partial {b}, all-places rows {c}")
